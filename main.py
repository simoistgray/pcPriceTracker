import pickle
import random

from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook, Workbook
from datetime import datetime
from openpyxl.utils import column_index_from_string
import smtplib
from email.mime.text import MIMEText
from selenium.webdriver.edge.options import Options


def send_sms_via_email(message):
    # Replace with your email credentials
    sender_email = "email@email.com"
    sender_password = "insertPassword"

    your_email = 'yourname@gmail.com'

    msg = MIMEText(message)
    msg["From"] = sender_email
    msg["To"] = your_email
    msg["Subject"] = "Sale Alert"

    # Sending the email
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, your_email, msg.as_string())


def getPrices(url):

    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ', 'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ', 'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EH', 'EI', 'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP', 'EQ', 'ER', 'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ', 'FA', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FI', 'FJ', 'FK', 'FL', 'FM', 'FN', 'FO', 'FP', 'FQ', 'FR', 'FS', 'FT', 'FU', 'FV', 'FW', 'FX', 'FY', 'FZ', 'GA', 'GB', 'GC', 'GD', 'GE', 'GF', 'GG', 'GH', 'GI', 'GJ', 'GK', 'GL', 'GM', 'GN', 'GO', 'GP', 'GQ', 'GR', 'GS', 'GT', 'GU', 'GV', 'GW', 'GX', 'GY', 'GZ', 'HA', 'HB', 'HC', 'HD', 'HE', 'HF', 'HG', 'HH', 'HI', 'HJ', 'HK', 'HL', 'HM', 'HN', 'HO', 'HP', 'HQ', 'HR', 'HS', 'HT', 'HU', 'HV', 'HW', 'HX', 'HY', 'HZ', 'IA', 'IB', 'IC', 'ID', 'IE', 'IF', 'IG', 'IH', 'II', 'IJ', 'IK', 'IL', 'IM', 'IN', 'IO', 'IP', 'IQ', 'IR', 'IS', 'IT', 'IU', 'IV', 'IW', 'IX', 'IY', 'IZ', 'JA', 'JB', 'JC', 'JD', 'JE', 'JF', 'JG', 'JH', 'JI', 'JJ', 'JK', 'JL', 'JM', 'JN', 'JO', 'JP', 'JQ', 'JR', 'JS', 'JT', 'JU', 'JV', 'JW', 'JX', 'JY', 'JZ', 'KA', 'KB', 'KC', 'KD', 'KE', 'KF', 'KG', 'KH', 'KI', 'KJ', 'KK', 'KL', 'KM', 'KN', 'KO', 'KP', 'KQ', 'KR', 'KS', 'KT', 'KU', 'KV', 'KW', 'KX', 'KY', 'KZ', 'LA', 'LB', 'LC', 'LD', 'LE', 'LF', 'LG', 'LH', 'LI', 'LJ', 'LK', 'LL', 'LM', 'LN', 'LO', 'LP', 'LQ', 'LR', 'LS', 'LT', 'LU', 'LV', 'LW', 'LX', 'LY', 'LZ', 'MA', 'MB', 'MC', 'MD', 'ME', 'MF', 'MG', 'MH', 'MI', 'MJ', 'MK', 'ML', 'MM', 'MN', 'MO', 'MP', 'MQ', 'MR', 'MS', 'MT', 'MU', 'MV', 'MW', 'MX', 'MY', 'MZ', 'NA', 'NB', 'NC', 'ND', 'NE', 'NF', 'NG', 'NH', 'NI', 'NJ', 'NK', 'NL', 'NM', 'NN', 'NO', 'NP', 'NQ', 'NR', 'NS', 'NT', 'NU', 'NV', 'NW', 'NX', 'NY', 'NZ', 'OA', 'OB', 'OC', 'OD', 'OE', 'OF', 'OG', 'OH', 'OI', 'OJ', 'OK', 'OL', 'OM', 'ON', 'OO', 'OP', 'OQ', 'OR', 'OS', 'OT', 'OU', 'OV', 'OW', 'OX', 'OY', 'OZ', 'PA', 'PB', 'PC', 'PD', 'PE', 'PF', 'PG', 'PH', 'PI', 'PJ', 'PK', 'PL', 'PM', 'PN', 'PO', 'PP', 'PQ', 'PR', 'PS', 'PT', 'PU', 'PV', 'PW', 'PX', 'PY', 'PZ', 'QA', 'QB', 'QC', 'QD', 'QE', 'QF', 'QG', 'QH', 'QI', 'QJ', 'QK', 'QL', 'QM', 'QN', 'QO', 'QP', 'QQ', 'QR', 'QS', 'QT', 'QU', 'QV', 'QW', 'QX', 'QY', 'QZ', 'RA', 'RB', 'RC', 'RD', 'RE', 'RF', 'RG', 'RH', 'RI', 'RJ', 'RK', 'RL', 'RM', 'RN', 'RO', 'RP', 'RQ', 'RR', 'RS', 'RT', 'RU', 'RV', 'RW', 'RX', 'RY', 'RZ', 'SA', 'SB', 'SC', 'SD', 'SE', 'SF', 'SG', 'SH', 'SI', 'SJ', 'SK', 'SL', 'SM', 'SN', 'SO', 'SP', 'SQ', 'SR', 'SS', 'ST', 'SU', 'SV', 'SW', 'SX', 'SY', 'SZ', 'TA', 'TB', 'TC', 'TD', 'TE', 'TF', 'TG', 'TH', 'TI', 'TJ', 'TK', 'TL', 'TM', 'TN', 'TO', 'TP', 'TQ', 'TR', 'TS', 'TT', 'TU', 'TV', 'TW', 'TX', 'TY', 'TZ', 'UA', 'UB', 'UC', 'UD', 'UE', 'UF', 'UG', 'UH', 'UI', 'UJ', 'UK', 'UL', 'UM', 'UN', 'UO', 'UP', 'UQ', 'UR', 'US', 'UT', 'UU', 'UV', 'UW', 'UX', 'UY', 'UZ', 'VA', 'VB', 'VC', 'VD', 'VE', 'VF', 'VG', 'VH', 'VI', 'VJ', 'VK', 'VL', 'VM', 'VN', 'VO', 'VP', 'VQ', 'VR', 'VS', 'VT', 'VU', 'VV', 'VW', 'VX', 'VY', 'VZ', 'WA', 'WB', 'WC', 'WD', 'WE', 'WF', 'WG', 'WH', 'WI', 'WJ', 'WK', 'WL', 'WM', 'WN', 'WO', 'WP', 'WQ', 'WR', 'WS', 'WT', 'WU', 'WV', 'WW', 'WX', 'WY', 'WZ', 'XA', 'XB', 'XC', 'XD', 'XE', 'XF', 'XG', 'XH', 'XI', 'XJ', 'XK', 'XL', 'XM', 'XN', 'XO', 'XP', 'XQ', 'XR', 'XS', 'XT', 'XU', 'XV', 'XW', 'XX', 'XY', 'XZ', 'YA', 'YB', 'YC', 'YD', 'YE', 'YF', 'YG', 'YH', 'YI', 'YJ', 'YK', 'YL', 'YM', 'YN', 'YO', 'YP', 'YQ', 'YR', 'YS', 'YT', 'YU', 'YV', 'YW', 'YX', 'YY', 'YZ', 'ZA', 'ZB', 'ZC', 'ZD', 'ZE', 'ZF', 'ZG', 'ZH', 'ZI', 'ZJ', 'ZK', 'ZL', 'ZM', 'ZN', 'ZO', 'ZP', 'ZQ', 'ZR', 'ZS', 'ZT', 'ZU', 'ZV', 'ZW', 'ZX', 'ZY', 'ZZ']

    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")

    driver = webdriver.Edge(options=options)

    driver.get(url)

    page_source = driver.page_source

    soup = BeautifulSoup(page_source, "html.parser")

    driver.quit()

    prices = []
    price_elements = soup.find_all('td', class_='td__price')

    for element in price_elements:
        price_tag = element.find('a')
        if price_tag:
            price_text = price_tag.get_text(strip=True)
            prices.append(price_text[1:])

    wb = load_workbook('path/to/excel/file')

    sheet = wb.active
    next_column = get_next_column(sheet, columns)

    now = datetime.now()
    formatted_date = now.strftime("%B %d, %Y, %H:%M:%S")

    if check_against_previous_prices(sheet, columns[columns.index(next_column) - 1], prices):
        sheet[next_column + str(2)] = formatted_date
        i = 3
        for price in prices:
            today_cell = "C" + str(i)
            date_cell = next_column + str(i)

            sheet["B" + str(i)] = (float(price) - float(sheet["G" + str(i)].value)) / float(sheet["G" + str(i)].value)
            sheet[today_cell] = float(price)
            sheet[date_cell] = float(price)

            wb.save('/Users/simongray/Desktop/pc build prices.xlsx')
            i += 1
    else:
        sheet[columns[columns.index(next_column) - 1] + str(2)] = formatted_date
        wb.save('/Users/simongray/Desktop/pc build prices.xlsx')


def get_next_column(sheet, columns):
    for col in columns:
        if sheet[col + str(2)].value is None:
            return col


def check_against_previous_prices(sheet, col, prices):
    prev_col = col
    x = 0
    for i in range(3, 11):
        if sheet[prev_col + str(i)].value != float(prices[x]):
            return True
        x += 1
    return False


def summarize_workbook(sheet, columns):
    toKeep = ['A', 'B', 'C']
    finishedColumn = 3
    for col in columns:
        if columns.index(col) >= finishedColumn:
            if sheet[col + str(2)].value is None:
                return toKeep
            else:
                currentIndex = columns.index(col)
                maxRange = len(columns) - currentIndex
                samePrices = True
                i = 0
                while samePrices is True and maxRange - 2 > i:
                    i += 1
                    for x in range(3, 11):
                        if samePrices:
                            if sheet[columns[currentIndex + i] + str(x)].value is not None:
                                if sheet[col + str(x)].value != sheet[columns[currentIndex + i] + str(x)].value:
                                    samePrices = False
                                    if col not in toKeep:
                                        toKeep.append(col)
                                    if columns[currentIndex + i] not in toKeep:
                                        toKeep.append(columns[currentIndex + i])
                                    finishedColumn = currentIndex + i


def copy_columns_to_new_workbook(sheet, to_keep):
    new_wb = Workbook()
    new_sheet = new_wb.active

    for new_col_idx, col_letter in enumerate(to_keep, start=1):
        col_idx = column_index_from_string(col_letter)
        for row_idx, cell in enumerate(sheet.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True), start=1):
            new_sheet.cell(row=row_idx, column=new_col_idx, value=cell[0])

    return new_wb


def checkPercentages():
    alreadyPurchased = [4, 10]
    wb = load_workbook('path/to/excel/file')
    sheet = wb.active
    message = 'https://pcpartpicker.com/list/v3KRQP\nThese items are over 15% off:'
    data = ''
    try:
        with open("message.pickle", "rb") as f:
            data = pickle.load(f)
    except:
        with open("message.pickle", "wb") as f:
            pickle.dump(message, f)
    for x in range(3, 11):
        # print(sheet['B' + str(x)].value)
        if sheet['B' + str(x)].value < -.15 and x not in alreadyPurchased:
            message += '\n' + sheet['A' + str(x)].value + ": " + str(round((sheet['B' + str(x)].value * 100), 2)) + "%. Was orignally $" + str(round(sheet['F' + str(x)].value, 2)) + " and is now $" + str(round(sheet['C' + str(x)].value, 2)) + "."
    if message != 'https://pcpartpicker.com/list/v3KRQP\nThese items are over 15% off:' and (message != data or random.randint(1, 4) >= 2):
        data = message
        with open("message.pickle", "wb") as f:
            pickle.dump(data, f)
        send_sms_via_email(message)
    now = datetime.now()
    if now.minute == 0:
        message2 = 'https://pcpartpicker.com/list/v3KRQP\nPrices for CPU and Video Card:\n'
        message2 += "CPU: " + str(sheet['C3'].value) + "\n"
        message2 += "Video Card: " + str(sheet['C8'].value)
        send_sms_via_email(message2)


getPrices("https://pcpartpicker.com/list/v3KRQP")
checkPercentages()
